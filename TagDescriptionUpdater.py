import json
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from pathlib import Path
import threading
from datetime import datetime
import os
import sys
import subprocess

class DescriptionUpdaterUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Tag Description Updater")
        self.root.geometry("600x800")
        self.root.configure(bg='#f0f0f0')
        
        self.excel_path = None
        self.json_path = None
        self.output_path = None
        self.alarm_filter_var = tk.BooleanVar(value=True)
        
        # Color palette
        self.accent_color = '#0078d4'
        self.success_color = '#28a745'
        self.info_color = '#5c6bc0'
        
        self.setup_styles()
        self.setup_ui()
    
    def setup_styles(self):
        """Configure custom styles"""
        style = ttk.Style()
        style.theme_use('clam')
        
        bg_color = '#f0f0f0'
        frame_bg = '#ffffff'
        
        style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), background=bg_color, foreground='#2c3e50')
        style.configure('Header.TLabel', font=('Segoe UI', 10, 'bold'), background=frame_bg)
        style.configure('Info.TLabel', font=('Segoe UI', 9), background=frame_bg, foreground='#555555')
        style.configure('Card.TFrame', background=frame_bg, relief='flat', borderwidth=1)
        style.configure('Card.TLabelframe', background=frame_bg, relief='solid', borderwidth=1)
        style.configure('Card.TLabelframe.Label', font=('Segoe UI', 10, 'bold'), foreground='#2c3e50', background=frame_bg)
    
    def setup_ui(self):
        """Setup the UI"""
        # Outer container
        outer_container = tk.Frame(self.root, bg='#f0f0f0')
        outer_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Main container
        main_container = tk.Frame(outer_container, bg='#f0f0f0')
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_container, text="📋 Tag Description Updater", style='Title.TLabel')
        title_label.pack(pady=(0, 15))
        
        # Top section: File Selection
        top_section = tk.Frame(main_container, bg='#f0f0f0')
        top_section.pack(fill=tk.X, pady=(0, 15))
        
        file_frame = ttk.LabelFrame(top_section, text="📁 File Selection", style='Card.TLabelframe', padding=15)
        file_frame.pack(fill=tk.X)
        
        # Excel file selection with filter on same row
        excel_row = tk.Frame(file_frame, bg='#ffffff')
        excel_row.pack(fill=tk.X, pady=5)
        
        self.excel_btn = tk.Button(excel_row, text="Select Excel File", command=self.browse_excel,
                                    bg=self.accent_color, fg='white', font=('Segoe UI', 9, 'bold'),
                                    relief='flat', padx=15, pady=8, cursor='hand2', width=18)
        self.excel_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.excel_var = tk.StringVar(value="No file selected")
        excel_label = ttk.Label(excel_row, textvariable=self.excel_var, style='Info.TLabel')
        excel_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Filter button frame on the right
        filter_frame = tk.Frame(excel_row, bg='#ffffff')
        filter_frame.pack(side=tk.RIGHT, padx=5, anchor=tk.N)
        
        alarm_label = ttk.Label(filter_frame, text="Filter", style='Header.TLabel')
        alarm_label.pack(anchor=tk.CENTER)
        
        alarm_check = tk.Button(filter_frame, text="✓ Only Is Alarm = True", command=self.toggle_alarm_filter,
                               bg=self.success_color, fg='white', font=('Segoe UI', 9, 'bold'),
                               relief='flat', padx=15, pady=5, cursor='hand2')
        alarm_check.pack(anchor=tk.CENTER, pady=(0, 0))
        self.alarm_check_btn = alarm_check
        
        # JSON file selection
        json_row = tk.Frame(file_frame, bg='#ffffff')
        json_row.pack(fill=tk.X, pady=5)
        
        self.json_btn = tk.Button(json_row, text="Select JSON File", command=self.browse_json,
                                   bg=self.info_color, fg='white', font=('Segoe UI', 9, 'bold'),
                                   relief='flat', padx=15, pady=8, cursor='hand2', width=18)
        self.json_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.json_var = tk.StringVar(value="No file selected")
        json_label = ttk.Label(json_row, textvariable=self.json_var, style='Info.TLabel')
        json_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Middle section: Info and Logs
        middle_section = tk.Frame(main_container, bg='#f0f0f0')
        middle_section.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Info Card
        info_frame = ttk.LabelFrame(middle_section, text="ℹ️ How It Works", style='Card.TLabelframe', padding=15)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        info_text = ("1. Select an Excel file containing tag descriptions\n"
                    "   (Columns: Data Block, Tag Name, Is Alarm, Description)\n\n"
                    "2. Select a JSON file with tags to update (Ignition export format)\n\n"
                    "3. Click 'Update Descriptions' to process\n\n"
                    "The tool will match tags by Data Block and Tag Name,\n"
                    "then update their descriptions in the JSON file.")
        
        info_label = ttk.Label(info_frame, text=info_text, style='Info.TLabel', justify=tk.LEFT)
        info_label.pack(anchor=tk.W)
        
        # Output Log Card (Full width)
        log_frame = ttk.LabelFrame(middle_section, text="📝 Output Log", style='Card.TLabelframe', padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        log_container = tk.Frame(log_frame, bg='#ffffff')
        log_container.pack(fill=tk.BOTH, expand=True)
        
        self.output_text = scrolledtext.ScrolledText(log_container, height=10, wrap=tk.WORD, 
                                                      bg='#f8f9fa', fg='#2c3e50', font=('Consolas', 9), 
                                                      relief='flat', padx=10, pady=10, state=tk.DISABLED)
        self.output_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(log_container, command=self.output_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.output_text.config(yscrollcommand=scrollbar.set)
        
        # Bottom section: Controls
        bottom_section = tk.Frame(main_container, bg='#f0f0f0')
        bottom_section.pack(fill=tk.X, pady=(10, 0))
        
        self.update_btn = tk.Button(bottom_section, text="🚀 Update Descriptions", command=self.run_update,
                                     bg=self.success_color, fg='white', font=('Segoe UI', 10, 'bold'),
                                     relief='flat', padx=20, pady=8, cursor='hand2',
                                     activebackground='#218838', activeforeground='white')
        self.update_btn.pack(padx=5)
    
    def browse_excel(self):
        """Browse for Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_path = file_path
            self.excel_var.set(Path(file_path).name)
            self.log(f"✓ Selected Excel: {Path(file_path).name}\n")
    
    def browse_json(self):
        """Browse for JSON file"""
        file_path = filedialog.askopenfilename(
            title="Select JSON File",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if file_path:
            self.json_path = file_path
            self.json_var.set(Path(file_path).name)
            self.log(f"✓ Selected JSON: {Path(file_path).name}\n")
    
    def log(self, message):
        """Add text to output log"""
        self.output_text.config(state=tk.NORMAL)
        self.output_text.insert(tk.END, message)
        self.output_text.see(tk.END)
        self.output_text.config(state=tk.DISABLED)
        self.root.update()
    
    def clear_output(self):
        """Clear output log"""
        self.output_text.config(state=tk.NORMAL)
        self.output_text.delete(1.0, tk.END)
        self.output_text.config(state=tk.DISABLED)
    
    def toggle_alarm_filter(self):
        """Toggle alarm filter and update button appearance"""
        current_state = self.alarm_filter_var.get()
        self.alarm_filter_var.set(not current_state)
        
        if self.alarm_filter_var.get():
            self.alarm_check_btn.config(bg=self.success_color, text="✓ Only Is Alarm = True")
        else:
            self.alarm_check_btn.config(bg='#6c757d', text="○ All Tags")
    
    def run_update(self):
        """Run the update process in a separate thread"""
        if not self.excel_path or not self.json_path:
            messagebox.showerror("Error", "Please select both Excel and JSON files")
            return
        
        # Ask user to select output file with custom filename
        default_filename = "tags_updated.json"
        output_file = filedialog.asksaveasfilename(
            title="Save Updated Tags As",
            initialdir=str(Path(self.json_path).parent),
            initialfile=default_filename,
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if not output_file:
            self.log("✗ Output file selection cancelled\n")
            return
        
        self.output_path = output_file
        self.update_btn.config(state=tk.DISABLED)
        thread = threading.Thread(target=self.update_descriptions)
        thread.start()
    
    def update_descriptions(self):
        """Main update logic"""
        try:
            self.log("=" * 60 + "\n")
            self.log("Tag Description Updater\n")
            self.log("=" * 60 + "\n\n")
            
            # Load descriptions from Excel
            self.log(f"1. Loading descriptions from {Path(self.excel_path).name}...\n")
            descriptions = self.load_descriptions_from_excel()
            self.log(f"   Found {len(descriptions)} tags with descriptions\n")
            if descriptions:
                first_entry = next(iter(descriptions.items()))
                self.log(f"   Sample entry: {first_entry[0]} -> Group: {first_entry[1].get('group')}\n\n")
            
            if len(descriptions) == 0:
                self.log("⚠ No matching descriptions found!\n")
                self.update_btn.config(state=tk.NORMAL)
                return
            
            # Load JSON
            self.log(f"2. Loading JSON from {Path(self.json_path).name}...\n")
            with open(self.json_path, 'r', encoding='utf-8') as f:
                tags_data = json.load(f)
            self.log("   JSON loaded successfully\n\n")
            
            # Update descriptions
            self.log(f"3. Updating tag descriptions...\n")
            updated = self.find_and_update_descriptions(tags_data.get("tags", []), descriptions)
            
            # Also check if root tag itself is in descriptions
            root_tag_name = tags_data.get("name")
            if root_tag_name and tags_data.get("tagType") == "UdtInstance":
                data_block_param = tags_data.get("parameters", {}).get("DataBlock", {}).get("value")
                if data_block_param:
                    lookup_key = f"{data_block_param}|{root_tag_name}"
                    if lookup_key in descriptions:
                        desc_data = descriptions[lookup_key]
                        group_value = desc_data.get('group') if isinstance(desc_data, dict) else None
                        if group_value:
                            if "parameters" not in tags_data:
                                tags_data["parameters"] = {}
                            tags_data["parameters"]["Group"] = {
                                "dataType": "String",
                                "value": group_value
                            }
            
            self.log(f"   Total updates: {updated}\n\n")
            
            # Save updated JSON
            self.log(f"4. Saving updated JSON to {Path(self.output_path).name}...\n")
            with open(self.output_path, 'w', encoding='utf-8') as f:
                json.dump(tags_data, f, indent=2, ensure_ascii=False)
            
            self.log(f"\n✓ Complete! Updated file saved as: {Path(self.output_path).name}\n")
            self.log("=" * 60 + "\n")
            
            messagebox.showinfo("Success", f"Updated {updated} tags successfully!\n\nOutput: {self.output_path}")
            
            # Open the output folder
            output_folder = str(Path(self.output_path).parent)
            if sys.platform == 'win32':
                os.startfile(output_folder)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', output_folder])
            else:
                subprocess.Popen(['xdg-open', output_folder])
            
        except Exception as e:
            self.log(f"\n✗ Error: {str(e)}\n")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        
        finally:
            self.update_btn.config(state=tk.NORMAL)
    
    def determine_substance(self, tag_name, origin, description):
        """Determine SUBSTANCE value based on keywords in tag name, origin, and description"""
        # Combine all text fields and convert to uppercase for case-insensitive matching
        search_text = f"{tag_name or ''} {origin or ''} {description or ''} ".upper()
        
        # Define keyword mappings - check in order of specificity
        substance_keywords = {
            'FUEL': ['FO', 'FUEL OIL', 'FUEL'],
            'WATER': ['FRESH WATER', 'FW'],
            'OIL': ['OIL'],
            'GRAY': ['GREY WATER', 'GREY', 'GW', 'GRAY'],
            'BLACK': ['G&B', 'GREY AND BLACK', 'BLACK WATER', 'BW', 'BLACK'],
            'SLUDGE': ['SLUDGE']
        }
        
        # Check for keyword matches (order matters for specificity)
        for substance, keywords in substance_keywords.items():
            for keyword in keywords:
                if keyword in search_text:
                    return substance
        
        return None
    
    def determine_substance_analog(self, tag_name, origin, description):
        """Determine SUBSTANCE value for ANALOG/ANL UDTs based on keywords"""
        # Combine all text fields and convert to uppercase for case-insensitive matching
        search_text = f"{tag_name or ''} {origin or ''} {description or ''} ".upper()
        
        # Define keyword mappings for ANALOG/ANL UDTs
        substance_keywords = {
            'COOLANT': ['COOLANT'],
            'OIL': ['OIL','LO','LUBE OIL'],
            'FUEL': ['FUEL','FO','FUEL OIL'],
            'WATER': ['WATER'],
            'GAS': ['EXHAUST']
        }
        
        # Check for keyword matches
        for substance, keywords in substance_keywords.items():
            for keyword in keywords:
                if keyword in search_text:
                    return substance
        
        return None
    
    def load_descriptions_from_excel(self):
        """Load descriptions from Excel file (all sheets except SCADA_SIGNAL), include sheet name as 'group'"""
        descriptions = {}
        wb = openpyxl.load_workbook(self.excel_path)
        filter_alarms = self.alarm_filter_var.get() if hasattr(self, 'alarm_filter_var') else True
        # Process all sheets except SCADA_SIGNAL
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() == "SCADA_SIGNAL":
                self.log(f"   Skipping sheet: {sheet_name}\n")
                continue
            self.log(f"   Processing sheet: {sheet_name}\n")
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                data_block = row[0]  # Column A: Data Block
                tag_name = row[1]  # Column B: Tag Name
                is_alarm = row[5] if len(row) > 5 else False  # Column F: Is Alarm
                alarm_priority = row[6] if len(row) > 6 else None  # Column G: Alarm Priority
                origin = row[8] if len(row) > 8 else None  # Column I: Origin
                description = row[9] if len(row) > 9 else None  # Column J: Description
                # Apply alarm filter if enabled
                if filter_alarms and not is_alarm:
                    continue
                if tag_name and description and data_block:
                    key = f"{data_block}|{tag_name}"
                    descriptions[key] = {
                        'description': description,
                        'origin': origin,
                        'tag_name': tag_name,
                        'data_block': data_block,
                        'alarm_priority': alarm_priority,
                        'is_alarm': is_alarm,
                        'group': sheet_name  # Add sheet name as group
                    }
        return descriptions
    
    def find_and_update_descriptions(self, tags_list, descriptions, parent_data_block=None):
        """Recursively find and update tag descriptions, and write group (sheet name) into tags"""
        updated_count = 0
        for tag in tags_list:
            current_data_block = parent_data_block
            # Track the current data block/folder
            if tag.get("tagType") == "Folder":
                current_data_block = tag.get("name")
            # For AtomicTags, check if they match Excel entries
            if tag.get("tagType") == "AtomicTag":
                tag_name = tag.get("name")
                lookup_key = f"{current_data_block}|{tag_name}"
                if lookup_key in descriptions:
                    desc_data = descriptions[lookup_key]
                    desc_value = desc_data['description'] if isinstance(desc_data, dict) else desc_data
                    tag["value"] = desc_value
                    # Write group (sheet name) into tag
                    if isinstance(desc_data, dict) and 'group' in desc_data:
                        tag["group"] = {
                            "dataType": "String",
                            "value": desc_data['group']
                        }
                    self.log(f"   ✓ {lookup_key} -> {desc_value}\n")
                    updated_count += 1
            # For UDT instances with alarm configuration (Configuration/Alarm_Description)
            if tag.get("tagType") == "UdtInstance":
                alarm_name = tag.get("name")
                # Get DataBlock from parameters if available
                data_block_param = tag.get("parameters", {}).get("DataBlock", {}).get("value")
                if data_block_param:
                    lookup_key = f"{data_block_param}|{alarm_name}"
                else:
                    # Fall back to parent folder name
                    lookup_key = f"{current_data_block}|{alarm_name}"
                if lookup_key in descriptions:
                    desc_data = descriptions[lookup_key]
                    description_text = desc_data['description'] if isinstance(desc_data, dict) else desc_data
                    origin_text = desc_data.get('origin') if isinstance(desc_data, dict) else None
                    tag_name = desc_data.get('tag_name') if isinstance(desc_data, dict) else alarm_name
                    data_block = desc_data.get('data_block') if isinstance(desc_data, dict) else current_data_block
                    alarm_priority = desc_data.get('alarm_priority') if isinstance(desc_data, dict) else None
                    is_alarm = desc_data.get('is_alarm') if isinstance(desc_data, dict) else False
                    group_value = desc_data.get('group') if isinstance(desc_data, dict) else None
                    # Write group (sheet name) into tag parameters
                    if group_value:
                        if "parameters" not in tag:
                            tag["parameters"] = {}
                        tag["parameters"]["Group"] = {
                            "dataType": "String",
                            "value": group_value
                        }
                    # Get UDT type name for checking ANALOG/ANL
                    udt_type = tag.get("type", "").upper()
                    is_analog_udt = "ANALOG" in udt_type or "ANL" in udt_type
                    if is_analog_udt:
                        self.log(f"      [DEBUG] Detected ANALOG UDT: {udt_type}\n")
                    # Only use Method 1: Find Configuration/Alarm_Description structure
                    config_found = False
                    for config_tag in tag.get("tags", []):
                        if config_tag.get("name") == "Configuration" and config_tag.get("tagType") == "Folder":
                            config_found = True
                            substance_updated = False
                            for sub_tag in config_tag.get("tags", []):
                                if sub_tag.get("name") == "Alarm_Description":
                                    sub_tag["value"] = description_text
                                    updated_count += 1
                                elif sub_tag.get("name") == "Label":
                                    sub_tag["value"] = description_text
                                    updated_count += 1
                                # Update Alarm_Priority if Is_Alarm is True
                                elif sub_tag.get("name") == "Alarm_Priority" and is_alarm:
                                    if alarm_priority is not None:
                                        sub_tag["value"] = alarm_priority
                                        self.log(f"      └─ Alarm_Priority: {alarm_priority}\n")
                                        updated_count += 1
                                # For TANKS data block, determine and set Substance
                                elif sub_tag.get("name") == "Substance" and data_block and data_block.upper() == "TANKS":
                                    substance = self.determine_substance(tag_name, origin_text, description_text)
                                    old_value = sub_tag.get("value", "")
                                    if substance:
                                        sub_tag["value"] = substance
                                        self.log(f"      └─ Substance (TANKS): '{old_value}' → '{substance}' ✓\n")
                                        updated_count += 1
                                        substance_updated = True
                                    else:
                                        self.log(f"      └─ Substance (TANKS): No keyword match (keeping '{old_value}')\n")
                                # For ANALOG/ANL UDTs, determine and set Substance
                                elif sub_tag.get("name") == "Substance" and is_analog_udt:
                                    substance = self.determine_substance_analog(tag_name, origin_text, description_text)
                                    old_value = sub_tag.get("value", "")
                                    if substance:
                                        sub_tag["value"] = substance
                                        self.log(f"      └─ Substance (ANALOG): '{old_value}' → '{substance}' ✓\n")
                                        updated_count += 1
                                        substance_updated = True
                                    else:
                                        self.log(f"      └─ Substance (ANALOG): No keyword match (keeping '{old_value}')\n")
                            # Log after updating fields
                            self.log(f"   ✓ {lookup_key} -> {description_text}\n")
                    if not config_found:
                        self.log(f"      [WARN] No 'Configuration/Folder' found in UDT structure\n")
                # Recursively process nested tags
                if len(tag.get("tags", [])) > 0:
                    updated_count += self.find_and_update_descriptions(
                        tag.get("tags", []), 
                        descriptions, 
                        current_data_block
                    )
            # Recursively process Folder tags
            elif tag.get("tagType") == "Folder" and "tags" in tag:
                updated_count += self.find_and_update_descriptions(
                    tag.get("tags", []), 
                    descriptions, 
                    current_data_block
                )
        return updated_count

def main():
    root = tk.Tk()
    app = DescriptionUpdaterUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
